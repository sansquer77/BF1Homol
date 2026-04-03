"""Operações de backup por tabela em Excel."""

from __future__ import annotations

import ast
import io
import json
from datetime import datetime
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter

from db.backup_utils import (
	_get_pk_columns,
	_get_serial_columns,
	_get_tables_with_fk_children,
	_list_tables,
	_quote_identifier,
)
from db.backup_validate import (
	_get_required_columns_for_insert,
	_get_table_column_types,
	_prevalidate_fk_values,
	_table_columns,
)
from db.db_schema import db_connect


def _normalize_excel_typed_value(value: Any, data_type: str) -> Any:
	dtype = (data_type or "").lower()
	if value is None:
		return None

	if dtype in {"json", "jsonb"}:
		if isinstance(value, (dict, list, tuple)):
			return json.dumps(value, ensure_ascii=False)
		if isinstance(value, str):
			raw = value.strip()
			if not raw:
				return None
			try:
				parsed = json.loads(raw)
				return json.dumps(parsed, ensure_ascii=False)
			except Exception:
				try:
					parsed = ast.literal_eval(raw)
					if isinstance(parsed, (dict, list, tuple)):
						return json.dumps(parsed, ensure_ascii=False)
				except Exception:
					return value
		return value

	if dtype == "array":
		if isinstance(value, (list, tuple)):
			return list(value)
		if isinstance(value, str):
			raw = value.strip()
			if not raw:
				return None
			try:
				parsed = ast.literal_eval(raw)
				if isinstance(parsed, (list, tuple)):
					return list(parsed)
			except Exception:
				return value
		return value

	return value


def _prepare_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
	if df.empty:
		return df

	safe_df = df.copy()
	for col in safe_df.columns:
		series = safe_df[col]
		if pd.api.types.is_datetime64tz_dtype(series.dtype):
			safe_df[col] = series.dt.tz_convert("UTC").dt.tz_localize(None)
			continue

		if pd.api.types.is_object_dtype(series.dtype):
			def _normalize_obj(value: Any) -> Any:
				if value is None:
					return None
				if isinstance(value, datetime):
					if value.tzinfo is None:
						return value
					return pd.Timestamp(value).tz_convert("UTC").tz_localize(None).to_pydatetime()
				return value

			safe_df[col] = series.map(_normalize_obj)

	return safe_df


def _apply_excel_datetime_format(
	writer: pd.ExcelWriter,
	sheet_name: str,
	df: pd.DataFrame,
	col_types: dict[str, str],
) -> None:
	if df.empty:
		return

	ws = writer.sheets.get(sheet_name)
	if ws is None:
		return

	datetime_cols: set[str] = set()
	for col in df.columns:
		dtype = df[col].dtype
		db_type = col_types.get(str(col).lower(), "")
		if pd.api.types.is_datetime64_any_dtype(dtype):
			datetime_cols.add(str(col))
			continue
		if "timestamp" in db_type or db_type in {"date", "time", "timetz"}:
			datetime_cols.add(str(col))

	if not datetime_cols:
		return

	excel_dt_format = "yyyy-mm-dd hh:mm:ss"
	excel_date_format = "yyyy-mm-dd"
	excel_time_format = "hh:mm:ss"

	for idx, col_name in enumerate(df.columns, start=1):
		if str(col_name) not in datetime_cols:
			continue

		db_type = col_types.get(str(col_name).lower(), "")
		if db_type == "date":
			number_format = excel_date_format
		elif db_type in {"time", "timetz"}:
			number_format = excel_time_format
		else:
			number_format = excel_dt_format

		col_letter = get_column_letter(idx)
		for row_idx in range(2, len(df) + 2):
			cell = ws[f"{col_letter}{row_idx}"]
			if cell.value is not None:
				cell.number_format = number_format


def download_tabela() -> None:
	tables = _list_tables()
	if not tables:
		st.info("No tables found for export.")
		return

	selected = st.selectbox("Table to export", tables, key="export_table_select")
	if not selected:
		return

	with db_connect() as conn:
		col_types = _get_table_column_types(conn, selected)
		c = conn.cursor()
		c.execute(f"SELECT * FROM {_quote_identifier(selected)}")
		rows = c.fetchall() or []
		col_names = [desc[0] for desc in c.description] if c.description else []

	df = pd.DataFrame([list(r.values()) for r in rows] if rows else [], columns=col_names)
	df_excel = _prepare_dataframe_for_excel(df)

	buffer = io.BytesIO()
	with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
		df_excel.to_excel(writer, index=False, sheet_name="data")
		_apply_excel_datetime_format(writer, "data", df_excel, col_types)
	buffer.seek(0)

	st.download_button(
		label=f"Download table {selected} (.xlsx)",
		data=buffer.getvalue(),
		file_name=f"{selected}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
		mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		on_click="ignore",
		width="stretch",
	)


def upload_tabela() -> None:
	tables = _list_tables()
	if not tables:
		st.info("No tables found for import.")
		return

	selected = st.selectbox("Destination table", tables, key="import_table_select")
	uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"], key="upload_table_xlsx")
	validate_fks = st.checkbox(
		"Pré-validar chaves estrangeiras antes de importar (recomendado)",
		value=True,
		key="import_table_validate_fks",
	)
	if not selected or not uploaded:
		return

	if st.button("Import table", type="primary", width="stretch"):
		df = pd.read_excel(uploaded)
		df.columns = [str(col).strip() for col in df.columns]
		db_cols = _table_columns(selected)
		use_cols = [c for c in df.columns if c in db_cols]
		if not use_cols:
			st.error("No compatible columns were found.")
			return

		with db_connect() as conn:
			col_types = _get_table_column_types(conn, selected)
			required_cols = _get_required_columns_for_insert(conn, selected)
			missing_required = [c for c in required_cols if c not in use_cols]
			if missing_required:
				st.error(
					"Importação bloqueada: o arquivo não contém colunas obrigatórias "
					f"da tabela '{selected}'."
				)
				st.caption(
					"Isso normalmente indica seleção incorreta da tabela de destino "
					"ou arquivo Excel de outra tabela."
				)
				st.caption(f"Colunas obrigatórias ausentes: {', '.join(missing_required)}")
				return

			payload = df[use_cols].astype(object)
			rows = []
			normalized_cells = 0
			for row in payload.itertuples(index=False, name=None):
				normalized_row: list[Any] = []
				for idx, value in enumerate(row):
					col_name = use_cols[idx]
					base_value = None if pd.isna(value) else value
					typed_value = _normalize_excel_typed_value(base_value, col_types.get(col_name.lower(), ""))
					if typed_value is not base_value:
						normalized_cells += 1
					normalized_row.append(typed_value)
				rows.append(tuple(normalized_row))

			col_idx = {col: idx for idx, col in enumerate(use_cols)}
			for req_col in required_cols:
				idx = col_idx.get(req_col)
				if idx is None:
					continue
				for row_number, row in enumerate(rows, start=2):
					if row[idx] is None:
						st.error(
							"Importação bloqueada: coluna obrigatória com valor vazio "
							f"na tabela '{selected}'."
						)
						st.caption(f"Coluna: {req_col} | Linha Excel: {row_number}")
						return

			fk_parent_tables = _get_tables_with_fk_children(conn)
			is_fk_parent = selected.lower() in {t.lower() for t in fk_parent_tables}

			if validate_fks:
				fk_errors = _prevalidate_fk_values(conn, selected, use_cols, rows)
				if fk_errors:
					st.error(
						"Importação bloqueada por inconsistência de FK no arquivo Excel. "
						"Corrija os valores e tente novamente."
					)
					for item in fk_errors:
						st.caption(f"- {item}")
					return

			c = conn.cursor()
			col_sql = ", ".join(_quote_identifier(col) for col in use_cols)
			placeholders = ", ".join(["%s"] * len(use_cols))

			if is_fk_parent:
				pk_cols = _get_pk_columns(conn, selected)
				if not pk_cols:
					st.error(
						f"Importação bloqueada: a tabela '{selected}' é referenciada por FK "
						"e não possui PRIMARY KEY detectada para UPSERT seguro."
					)
					st.info(
						"Para evitar quebra de integridade, esse cenário não executa TRUNCATE CASCADE. "
						"Defina uma PK na tabela ou use restore SQL completo."
					)
					return

				pk_set = {col.lower() for col in pk_cols}
				update_cols = [col for col in use_cols if col.lower() not in pk_set]
				conflict_target = ", ".join(_quote_identifier(col) for col in pk_cols)

				if update_cols:
					update_clause = ", ".join(
						f"{_quote_identifier(col)} = EXCLUDED.{_quote_identifier(col)}" for col in update_cols
					)
					upsert_sql = (
						f"INSERT INTO {_quote_identifier(selected)} ({col_sql}) "
						f"VALUES ({placeholders}) "
						f"ON CONFLICT ({conflict_target}) DO UPDATE SET {update_clause}"
					)
				else:
					upsert_sql = (
						f"INSERT INTO {_quote_identifier(selected)} ({col_sql}) "
						f"VALUES ({placeholders}) "
						f"ON CONFLICT ({conflict_target}) DO NOTHING"
					)

				c.executemany(upsert_sql, rows)
				st.info(
					f"⚠️ '{selected}' é referenciada por outras tabelas: linhas existentes foram "
					"atualizadas (UPSERT) e linhas ausentes no Excel foram mantidas. "
					"Nenhum dado filho foi apagado."
				)
			else:
				c.execute(f"TRUNCATE TABLE {_quote_identifier(selected)} RESTART IDENTITY CASCADE")
				c.executemany(
					f"INSERT INTO {_quote_identifier(selected)} ({col_sql}) VALUES ({placeholders})",
					rows,
				)

			serial_cols = _get_serial_columns(conn, selected)
			for col in serial_cols:
				qt = _quote_identifier(selected)
				qc = _quote_identifier(col)
				c.execute(
					f"""
					SELECT setval(
						pg_get_serial_sequence(%s, %s),
						COALESCE((SELECT MAX({qc}) FROM {qt}), 1),
						true
					)
					""",
					(selected, col),
				)
			conn.commit()

		if normalized_cells > 0:
			st.info(
				f"{normalized_cells} valores foram normalizados para tipos PostgreSQL "
				"(JSON/ARRAY) durante a importação."
			)

		st.success(f"Table {selected} imported successfully.")

__all__ = ["download_tabela", "upload_tabela"]
